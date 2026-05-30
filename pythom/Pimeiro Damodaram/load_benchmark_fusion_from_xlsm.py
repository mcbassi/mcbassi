#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import math
import os
import re
import sys
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
import pymysql
from openpyxl.utils import get_column_letter
from pymysql.cursors import DictCursor


RAW_BATCH_SIZE = 1000
ISSUE_BATCH_SIZE = 500

DAMODARAN_SHEETS = {
    "Damodaran 2021": 2021,
    "Damodaran 2022": 2022,
    "Damodaran 2023": 2023,
    "Damodaran 2024": 2024,
    "Damodaran 2024 (2)": 2024,
}

LOCAL_CIIU_SHEETS = {
    "2019": 2019,
    "2020": 2020,
}

DUPLICATE_CURATED_SHEETS = {"Damodaran 2024 (2)"}

# Physical column -> metric_code, based on the schema seed.
DAMODARAN_COL_TO_METRIC = {
    3: "capex_usd_mm",
    4: "depreciation_amortization_usd_mm",
    5: "capex_to_depreciation",
    6: "acquisitions_usd_mm",
    7: "net_rd_usd_mm",
    8: "net_capex_to_sales",
    9: "net_capex_to_ebit_after_tax",
    10: "sales_to_invested_capital",
    14: "beta",
    15: "de_ratio",
    16: "effective_tax_rate",
    17: "unlevered_beta",
    18: "cash_to_firm_value",
    19: "unlevered_beta_cash_adjusted",
    20: "hilo_risk",
    21: "stddev_equity",
    22: "stddev_operating_income_10y",
    31: "avg_unlevered_beta",
    32: "avg_levered_beta",
    33: "avg_correlation_market",
    34: "total_unlevered_beta",
    35: "total_levered_beta",
    39: "ev_to_ebitdarnd_positive_ebitda",
    40: "ev_to_ebitda_positive_ebitda",
    41: "ev_to_ebit_positive_ebitda",
    42: "ev_to_ebit_after_tax_positive_ebitda",
    43: "ev_to_ebitdarnd_all_firms",
    44: "ev_to_ebitda_all_firms",
    45: "ev_to_ebit_all_firms",
    46: "ev_to_ebit_after_tax_all_firms",
    50: "price_to_sales",
    51: "net_margin",
    52: "ev_to_sales",
    53: "pretax_operating_margin",
    57: "gross_margin",
    59: "pretax_pre_stock_comp_operating_margin",
    60: "pretax_unadjusted_operating_margin",
    61: "aftertax_unadjusted_operating_margin",
    62: "pretax_lease_adjusted_margin",
    63: "aftertax_lease_adjusted_margin",
    64: "pretax_lease_rd_adjusted_margin",
    65: "aftertax_lease_rd_adjusted_margin",
    66: "ebitda_to_sales",
    67: "ebitda_sga_to_sales",
    68: "ebitda_rd_to_sales",
    69: "cogs_to_sales",
    70: "rd_to_sales",
    71: "sga_to_sales",
    72: "stock_comp_to_sales",
    73: "lease_expense_to_sales",
    # 77 is a repeated Beta in the value creation block; skipped intentionally.
    78: "roe",
    79: "cost_of_equity",
    80: "roe_minus_coe",
    81: "bv_of_equity_usd_mm",
    82: "equity_eva_usd_mm",
    83: "roc",
    84: "cost_of_capital",
    85: "roc_minus_wacc",
    86: "bv_of_capital_usd_mm",
    87: "eva_usd_mm",
    91: "accounts_receivable_to_sales",
    92: "inventory_to_sales",
    93: "accounts_payable_to_sales",
    94: "noncash_working_capital_to_sales",
}

DAMODARAN_BETA_HISTORY_COLS = {
    23: 2021,
    24: 2022,
    25: 2023,
    26: 2024,
}
DAMODARAN_BETA_AVG_COL = 27


class ImporterError(Exception):
    pass


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def detect_sheet_role(sheet_name: str) -> str:
    if sheet_name in DUPLICATE_CURATED_SHEETS:
        return "DUPLICATE"
    if sheet_name in DAMODARAN_SHEETS:
        return "DAMODARAN"
    if sheet_name in LOCAL_CIIU_SHEETS:
        return "LOCAL_CIIU"
    return "OTHER"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def connect_mysql(args: argparse.Namespace):
    return pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database,
        charset="utf8mb4",
        cursorclass=DictCursor,
        autocommit=False,
    )


def load_metric_maps(conn) -> Tuple[Dict[str, int], Dict[str, int]]:
    sql = """
        SELECT ds.dataset_code, dm.metric_code, dm.metric_name, dm.id
        FROM dim_metric dm
        JOIN dim_dataset ds ON ds.id = dm.dataset_id
    """
    with conn.cursor() as cur:
        cur.execute(sql)
        rows = cur.fetchall()

    metric_id_by_code: Dict[str, int] = {}
    local_metric_id_by_header: Dict[str, int] = {}

    for row in rows:
        metric_id_by_code[row["metric_code"]] = row["id"]
        if row["dataset_code"] == "LOCAL_CIIU":
            local_metric_id_by_header[normalize_text(row["metric_name"])] = row["id"]

    required_codes = set(DAMODARAN_COL_TO_METRIC.values())
    missing = sorted(code for code in required_codes if code not in metric_id_by_code)
    if missing:
        raise ImporterError(
            "These metric_code values are missing in dim_metric: " + ", ".join(missing)
        )

    return metric_id_by_code, local_metric_id_by_header


def insert_import_batch(conn, path: Path, checksum: str, args: argparse.Namespace) -> int:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id FROM etl_import_batch WHERE source_checksum_sha256 = %s",
            (checksum,),
        )
        existing = cur.fetchone()
        if existing:
            raise ImporterError(
                f"This workbook checksum is already loaded in etl_import_batch.id={existing['id']}. "
                "If you really want another raw batch, change the file or clear that row first."
            )

        cur.execute(
            """
            INSERT INTO etl_import_batch (
                source_file_name,
                source_file_type,
                source_checksum_sha256,
                source_origin,
                imported_by,
                notes
            ) VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                path.name,
                path.suffix.lstrip(".").lower() or "xlsm",
                checksum,
                args.source_origin,
                args.imported_by,
                args.notes,
            ),
        )
        return int(cur.lastrowid)


def upsert_dim_damodaran_industry(conn, industry_name: str) -> int:
    normalized = normalize_text(industry_name)
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO dim_damodaran_industry (
                industry_name,
                industry_name_normalized,
                is_total_market,
                is_financial_sector,
                active
            ) VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                industry_name = VALUES(industry_name),
                active = 1,
                id = LAST_INSERT_ID(id)
            """,
            (industry_name, normalized, 0, 0, 1),
        )
        return int(cur.lastrowid)


def upsert_dim_ciiu(conn, ciiu_code: str, ciiu_description: Optional[str]) -> int:
    section = ciiu_code[:1] if ciiu_code else None
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO dim_ciiu (
                ciiu_code,
                ciiu_description,
                ciiu_section_letter
            ) VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE
                ciiu_description = COALESCE(VALUES(ciiu_description), ciiu_description),
                ciiu_section_letter = COALESCE(VALUES(ciiu_section_letter), ciiu_section_letter),
                id = LAST_INSERT_ID(id)
            """,
            (ciiu_code, ciiu_description, section),
        )
        return int(cur.lastrowid)


def parse_decimal(value: Any) -> Tuple[Optional[Decimal], Optional[str], Optional[str]]:
    """
    Returns: (decimal_value, text_fallback, issue_type)
    issue_type examples: text_number, formula_error, special_text
    """
    if value is None:
        return None, None, None

    if isinstance(value, bool):
        return Decimal(1 if value else 0), None, None

    if isinstance(value, Decimal):
        return value, None, None

    if isinstance(value, int):
        return Decimal(value), None, None

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None, str(value), "special_text"
        return Decimal(str(value)), None, None

    raw = str(value).strip()
    if raw == "":
        return None, None, None

    upper = raw.upper()
    if upper in {"NA", "N/A", "NM", "NULL", "NONE"}:
        return None, raw, "special_text"
    if raw.startswith("#"):
        return None, raw, "formula_error"

    cleaned = raw.replace("$", "").replace("€", "").replace("£", "")
    cleaned = cleaned.replace(" ", "")
    cleaned = re.sub(r"[^0-9,\.\-]", "", cleaned)

    if cleaned in {"", "-", ".", ",", "-.", "-,"}:
        return None, raw, "special_text"

    # Heuristics for broken strings like "$1,441,58".
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif cleaned.count(",") > 1 and "." not in cleaned:
        parts = cleaned.split(",")
        cleaned = "".join(parts[:-1]) + "." + parts[-1]
    elif cleaned.count(".") > 1 and "," not in cleaned:
        parts = cleaned.split(".")
        cleaned = "".join(parts[:-1]) + "." + parts[-1]
    elif cleaned.count(",") == 1 and "." not in cleaned:
        left, right = cleaned.split(",")
        if len(right) in {1, 2}:
            cleaned = f"{left}.{right}"
        else:
            cleaned = left + right

    try:
        return Decimal(cleaned), None, ("text_number" if raw != cleaned else None)
    except InvalidOperation:
        return None, raw, "special_text"


def excel_cell_payload(formula_cell, value_cell) -> Tuple[str, Optional[str], Optional[Decimal], Optional[datetime], Optional[str]]:
    formula_text: Optional[str] = None
    raw_text: Optional[str] = None
    raw_decimal: Optional[Decimal] = None
    raw_date: Optional[datetime] = None

    if formula_cell.data_type == "f" or (
        isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
    ):
        formula_text = str(formula_cell.value) if formula_cell.value is not None else None

    value = value_cell.value
    if value is None:
        return "blank", None, None, None, formula_text

    if value_cell.is_date and isinstance(value, datetime):
        return "date", None, None, value, formula_text

    if isinstance(value, bool):
        raw_text = "1" if value else "0"
        return "bool", raw_text, None, None, formula_text

    if isinstance(value, (int, float, Decimal)):
        parsed, text_fallback, _ = parse_decimal(value)
        return "number", text_fallback, parsed, None, formula_text

    if formula_text:
        parsed, text_fallback, issue_type = parse_decimal(value)
        if parsed is not None:
            return "number", text_fallback, parsed, None, formula_text
        return ("error" if issue_type == "formula_error" else "formula"), text_fallback or str(value), None, None, formula_text

    parsed, text_fallback, issue_type = parse_decimal(value)
    if parsed is not None:
        return "number", text_fallback, parsed, None, formula_text

    return ("error" if issue_type == "formula_error" else "text"), str(value), None, None, formula_text


def insert_issue_rows(conn, issue_rows: List[Tuple[Any, ...]]) -> None:
    if not issue_rows:
        return
    with conn.cursor() as cur:
        cur.executemany(
            """
            INSERT INTO etl_import_issue (
                import_batch_id,
                severity,
                issue_type,
                sheet_name,
                row_num,
                col_num,
                issue_message,
                raw_value_text
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            issue_rows,
        )


def import_raw_layer(conn, batch_id: int, formula_wb, value_wb) -> Dict[str, int]:
    issue_rows: List[Tuple[Any, ...]] = []
    import_sheet_ids: Dict[str, int] = {}

    for sheet_index, sheet_name in enumerate(formula_wb.sheetnames, start=1):
        formula_ws = formula_wb[sheet_name]
        value_ws = value_wb[sheet_name]
        role = detect_sheet_role(sheet_name)

        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO etl_import_sheet (
                    import_batch_id,
                    sheet_name,
                    sheet_index,
                    row_count,
                    column_count,
                    logical_role
                ) VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (batch_id, sheet_name, sheet_index, formula_ws.max_row, formula_ws.max_column, role),
            )
            import_sheet_id = int(cur.lastrowid)
        import_sheet_ids[sheet_name] = import_sheet_id

        if role == "DUPLICATE":
            issue_rows.append(
                (
                    batch_id,
                    "warning",
                    "duplicate_sheet",
                    sheet_name,
                    None,
                    None,
                    "Sheet marked as duplicate; raw cells imported, curated metrics will be skipped.",
                    None,
                )
            )

        batch_rows: List[Tuple[Any, ...]] = []
        for r in range(1, formula_ws.max_row + 1):
            for c in range(1, formula_ws.max_column + 1):
                fcell = formula_ws.cell(r, c)
                vcell = value_ws.cell(r, c)
                value_type, raw_text, raw_decimal, raw_date, formula_text = excel_cell_payload(fcell, vcell)
                batch_rows.append(
                    (
                        import_sheet_id,
                        r,
                        c,
                        f"{get_column_letter(c)}{r}",
                        value_type,
                        raw_text,
                        raw_decimal,
                        raw_date,
                        formula_text,
                    )
                )
                if len(batch_rows) >= RAW_BATCH_SIZE:
                    with conn.cursor() as cur:
                        cur.executemany(
                            """
                            INSERT INTO etl_import_cell (
                                import_sheet_id,
                                row_num,
                                col_num,
                                cell_ref,
                                value_type,
                                raw_value_text,
                                raw_value_decimal,
                                raw_value_date,
                                formula_text
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            batch_rows,
                        )
                    batch_rows.clear()
        if batch_rows:
            with conn.cursor() as cur:
                cur.executemany(
                    """
                    INSERT INTO etl_import_cell (
                        import_sheet_id,
                        row_num,
                        col_num,
                        cell_ref,
                        value_type,
                        raw_value_text,
                        raw_value_decimal,
                        raw_value_date,
                        formula_text
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    batch_rows,
                )

    if issue_rows:
        insert_issue_rows(conn, issue_rows)
    return import_sheet_ids


def upsert_damodaran_snapshot(conn, batch_id: int, asof_year: int, industry_id: int, number_of_firms: Optional[int], sheet_name: str, row_num: int) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO fact_damodaran_snapshot (
                import_batch_id,
                asof_year,
                damodaran_industry_id,
                number_of_firms,
                source_sheet_name,
                source_row_num
            ) VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                import_batch_id = VALUES(import_batch_id),
                number_of_firms = VALUES(number_of_firms),
                source_sheet_name = VALUES(source_sheet_name),
                source_row_num = VALUES(source_row_num),
                id = LAST_INSERT_ID(id)
            """,
            (batch_id, asof_year, industry_id, number_of_firms, sheet_name, row_num),
        )
        snapshot_id = int(cur.lastrowid)
        cur.execute("DELETE FROM fact_damodaran_metric WHERE snapshot_id = %s", (snapshot_id,))
        cur.execute("DELETE FROM fact_damodaran_beta_history WHERE snapshot_id = %s", (snapshot_id,))
        return snapshot_id


def upsert_local_snapshot(conn, batch_id: int, ref_year: int, ciiu_id: int, sector_label: Optional[str], percentile_value: Decimal, n_observations: Optional[int], sheet_name: str, row_num: int) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO fact_local_ciiu_percentile_snapshot (
                import_batch_id,
                ref_year,
                ciiu_id,
                sector_label,
                percentile_value,
                n_observations,
                source_sheet_name,
                source_row_num
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                import_batch_id = VALUES(import_batch_id),
                sector_label = VALUES(sector_label),
                n_observations = VALUES(n_observations),
                source_sheet_name = VALUES(source_sheet_name),
                source_row_num = VALUES(source_row_num),
                id = LAST_INSERT_ID(id)
            """,
            (batch_id, ref_year, ciiu_id, sector_label, percentile_value, n_observations, sheet_name, row_num),
        )
        snapshot_id = int(cur.lastrowid)
        cur.execute("DELETE FROM fact_local_ciiu_percentile_metric WHERE snapshot_id = %s", (snapshot_id,))
        return snapshot_id


def import_damodaran_curated(conn, batch_id: int, value_wb, metric_id_by_code: Dict[str, int]) -> None:
    issue_rows: List[Tuple[Any, ...]] = []

    for sheet_name, asof_year in DAMODARAN_SHEETS.items():
        if sheet_name in DUPLICATE_CURATED_SHEETS:
            continue

        ws = value_wb[sheet_name]
        for row_num in range(2, ws.max_row + 1):
            industry_name = ws.cell(row_num, 1).value
            if industry_name is None or str(industry_name).strip() == "":
                continue

            industry_name = str(industry_name).strip()
            industry_id = upsert_dim_damodaran_industry(conn, industry_name)

            firms_raw = ws.cell(row_num, 2).value
            firms_decimal, _, _ = parse_decimal(firms_raw)
            number_of_firms = int(firms_decimal) if firms_decimal is not None else None

            snapshot_id = upsert_damodaran_snapshot(
                conn=conn,
                batch_id=batch_id,
                asof_year=asof_year,
                industry_id=industry_id,
                number_of_firms=number_of_firms,
                sheet_name=sheet_name,
                row_num=row_num,
            )

            metric_rows: List[Tuple[Any, ...]] = []
            for col_num, metric_code in DAMODARAN_COL_TO_METRIC.items():
                metric_id = metric_id_by_code[metric_code]
                cell_value = ws.cell(row_num, col_num).value
                metric_value, metric_text, issue_type = parse_decimal(cell_value)
                metric_rows.append(
                    (
                        snapshot_id,
                        metric_id,
                        metric_value,
                        metric_text,
                        col_num,
                        str(ws.cell(1, col_num).value or ws.cell(2, col_num).value or ""),
                    )
                )
                if issue_type and metric_text:
                    issue_rows.append(
                        (
                            batch_id,
                            "warning",
                            issue_type,
                            sheet_name,
                            row_num,
                            col_num,
                            f"Damodaran metric '{metric_code}' imported as text/exception.",
                            metric_text,
                        )
                    )

            with conn.cursor() as cur:
                cur.executemany(
                    """
                    INSERT INTO fact_damodaran_metric (
                        snapshot_id,
                        metric_id,
                        metric_value,
                        metric_value_text,
                        source_col_num,
                        source_col_name
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    metric_rows,
                )

            beta_rows: List[Tuple[Any, ...]] = []
            for col_num, history_year in DAMODARAN_BETA_HISTORY_COLS.items():
                value = ws.cell(row_num, col_num).value
                metric_value, metric_text, issue_type = parse_decimal(value)
                beta_rows.append(
                    (snapshot_id, history_year, metric_value, 0, col_num)
                )
                if issue_type and metric_text:
                    issue_rows.append(
                        (
                            batch_id,
                            "warning",
                            issue_type,
                            sheet_name,
                            row_num,
                            col_num,
                            f"Damodaran beta history for {history_year} imported as text/exception.",
                            metric_text,
                        )
                    )

            avg_value, avg_text, avg_issue = parse_decimal(ws.cell(row_num, DAMODARAN_BETA_AVG_COL).value)
            beta_rows.append((snapshot_id, 0, avg_value, 1, DAMODARAN_BETA_AVG_COL))
            if avg_issue and avg_text:
                issue_rows.append(
                    (
                        batch_id,
                        "warning",
                        avg_issue,
                        sheet_name,
                        row_num,
                        DAMODARAN_BETA_AVG_COL,
                        "Damodaran average beta history imported as text/exception.",
                        avg_text,
                    )
                )

            with conn.cursor() as cur:
                cur.executemany(
                    """
                    INSERT INTO fact_damodaran_beta_history (
                        snapshot_id,
                        history_year,
                        beta_value,
                        is_average_row,
                        source_col_num
                    ) VALUES (%s, %s, %s, %s, %s)
                    """,
                    beta_rows,
                )

    if issue_rows:
        for i in range(0, len(issue_rows), ISSUE_BATCH_SIZE):
            insert_issue_rows(conn, issue_rows[i:i + ISSUE_BATCH_SIZE])


def import_local_curated(conn, batch_id: int, value_wb, local_metric_id_by_header: Dict[str, int]) -> None:
    issue_rows: List[Tuple[Any, ...]] = []

    for sheet_name, ref_year in LOCAL_CIIU_SHEETS.items():
        ws = value_wb[sheet_name]
        header_map: Dict[int, int] = {}
        for col_num in range(6, ws.max_column + 1):
            header = ws.cell(1, col_num).value
            normalized = normalize_text(header)
            metric_id = local_metric_id_by_header.get(normalized)
            if metric_id is None:
                issue_rows.append(
                    (
                        batch_id,
                        "warning",
                        "unmapped_header",
                        sheet_name,
                        1,
                        col_num,
                        "Local CIIU metric header not found in dim_metric.",
                        str(header),
                    )
                )
                continue
            header_map[col_num] = metric_id

        last_sector_label: Optional[str] = None
        for row_num in range(2, ws.max_row + 1):
            ciiu_code_raw = ws.cell(row_num, 2).value
            if ciiu_code_raw is None or str(ciiu_code_raw).strip() == "":
                continue

            ciiu_code = str(ciiu_code_raw).strip()
            sector_label = ws.cell(row_num, 3).value
            if sector_label is not None and str(sector_label).strip() != "":
                last_sector_label = str(sector_label).strip()
            else:
                sector_label = last_sector_label

            ciiu_id = upsert_dim_ciiu(conn, ciiu_code, sector_label)

            percentile_raw = ws.cell(row_num, 5).value
            percentile_value, percentile_text, percentile_issue = parse_decimal(percentile_raw)
            if percentile_value is None:
                issue_rows.append(
                    (
                        batch_id,
                        "error",
                        percentile_issue or "parse_error",
                        sheet_name,
                        row_num,
                        5,
                        "Percentile could not be parsed; row skipped.",
                        percentile_text or str(percentile_raw),
                    )
                )
                continue

            n_raw = ws.cell(row_num, 4).value
            n_value, _, _ = parse_decimal(n_raw)
            n_observations = int(n_value) if n_value is not None else None

            snapshot_id = upsert_local_snapshot(
                conn=conn,
                batch_id=batch_id,
                ref_year=ref_year,
                ciiu_id=ciiu_id,
                sector_label=sector_label,
                percentile_value=percentile_value,
                n_observations=n_observations,
                sheet_name=sheet_name,
                row_num=row_num,
            )

            metric_rows: List[Tuple[Any, ...]] = []
            for col_num, metric_id in header_map.items():
                cell_value = ws.cell(row_num, col_num).value
                metric_value, metric_text, issue_type = parse_decimal(cell_value)
                metric_rows.append(
                    (
                        snapshot_id,
                        metric_id,
                        metric_value,
                        metric_text,
                        col_num,
                        str(ws.cell(1, col_num).value or ""),
                    )
                )
                if issue_type and metric_text:
                    issue_rows.append(
                        (
                            batch_id,
                            "warning",
                            issue_type,
                            sheet_name,
                            row_num,
                            col_num,
                            "Local CIIU metric imported as text/exception.",
                            metric_text,
                        )
                    )

            with conn.cursor() as cur:
                cur.executemany(
                    """
                    INSERT INTO fact_local_ciiu_percentile_metric (
                        snapshot_id,
                        metric_id,
                        metric_value,
                        metric_value_text,
                        source_col_num,
                        source_col_name
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    metric_rows,
                )

    if issue_rows:
        for i in range(0, len(issue_rows), ISSUE_BATCH_SIZE):
            insert_issue_rows(conn, issue_rows[i:i + ISSUE_BATCH_SIZE])


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load the Damodaran workbook into the benchmark_fusion schema."
    )
    parser.add_argument("--file", required=True, help="Path to Datos procesados Damodaran.xlsm")
    parser.add_argument("--host", default=os.getenv("MYSQL_HOST", "127.0.0.1"))
    parser.add_argument("--port", type=int, default=int(os.getenv("MYSQL_PORT", "3306")))
    parser.add_argument("--user", default=os.getenv("MYSQL_USER", "root"))
    parser.add_argument("--password", default=os.getenv("MYSQL_PASSWORD", ""))
    parser.add_argument("--database", default=os.getenv("MYSQL_DATABASE", "benchmark_fusion"))
    parser.add_argument("--source-origin", default="manual_upload")
    parser.add_argument("--imported-by", default=os.getenv("USERNAME") or os.getenv("USER") or "manual")
    parser.add_argument("--notes", default="Carga inicial do workbook Damodaran/CIIU")
    parser.add_argument("--skip-raw", action="store_true", help="Skip etl_import_* raw layer.")
    parser.add_argument("--skip-curated", action="store_true", help="Skip curated fact tables.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.file).expanduser().resolve()
    if not workbook_path.exists():
        print(f"Arquivo não encontrado: {workbook_path}", file=sys.stderr)
        return 1

    checksum = sha256_file(workbook_path)

    formula_wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    value_wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)

    conn = connect_mysql(args)
    try:
        metric_id_by_code, local_metric_id_by_header = load_metric_maps(conn)
        batch_id = insert_import_batch(conn, workbook_path, checksum, args)

        if not args.skip_raw:
            import_raw_layer(conn, batch_id, formula_wb, value_wb)

        if not args.skip_curated:
            import_damodaran_curated(conn, batch_id, value_wb, metric_id_by_code)
            import_local_curated(conn, batch_id, value_wb, local_metric_id_by_header)

        conn.commit()
        print(f"Carga concluída com sucesso. etl_import_batch.id = {batch_id}")
        return 0
    except Exception as exc:
        conn.rollback()
        print(f"Erro na carga: {exc}", file=sys.stderr)
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
